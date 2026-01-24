"""
Excel Export Generator

Creates reproducible Excel files with 3 sheets:
1. Applications - One row per job application
2. Events - Timeline of all updates
3. ActionQueue - Pending actions requiring attention
"""

from typing import List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import Application, ApplicationEvent, ExportLog


class ExcelExporter:
    """Generate Excel reports from database"""

    def __init__(self, db: Session, user_id: int = None):
        self.db = db
        self.user_id = user_id

    def export_all(self, file_path: str) -> str:
        """
        Export all data to Excel file.

        Returns:
            file_path where Excel was saved
        """
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        self._create_applications_sheet(wb)
        self._create_events_sheet(wb)
        self._create_action_queue_sheet(wb)

        # Save file
        wb.save(file_path)

        # Log export
        self._log_export(file_path, wb)

        return file_path

    def _create_applications_sheet(self, wb):
        """Sheet 1: Applications"""
        ws = wb.create_sheet("Applications", 0)

        # Headers
        headers = [
            "ID", "Company", "Job Title", "Location", "Status",
            "Application Date", "Latest Email", "Action Required",
            "Action Type", "Action Deadline", "Confidence", "Events", "Links"
        ]

        self._write_headers(ws, headers)

        # Data - filter by user if specified
        query = self.db.query(Application)
        if self.user_id:
            query = query.filter(Application.user_id == self.user_id)
        applications = query.order_by(Application.created_at.desc()).all()

        for row_idx, app in enumerate(applications, start=2):
            ws.cell(row_idx, 1, app.id)
            ws.cell(row_idx, 2, app.company_name)
            ws.cell(row_idx, 3, app.job_title)
            ws.cell(row_idx, 4, app.location or "")
            ws.cell(row_idx, 5, app.current_status.value)
            ws.cell(row_idx, 6, app.application_date.strftime("%Y-%m-%d") if app.application_date else "")
            ws.cell(row_idx, 7, app.latest_email_date.strftime("%Y-%m-%d") if app.latest_email_date else "")
            ws.cell(row_idx, 8, "Yes" if app.action_required else "No")
            ws.cell(row_idx, 9, app.action_type.value if app.action_type else "")
            ws.cell(row_idx, 10, app.action_deadline.strftime("%Y-%m-%d") if app.action_deadline else "")
            ws.cell(row_idx, 11, f"{app.overall_confidence:.2f}" if app.overall_confidence else "")
            ws.cell(row_idx, 12, app.event_count)
            ws.cell(row_idx, 13, app.link_count)

        self._autosize_columns(ws)

    def _create_events_sheet(self, wb):
        """Sheet 2: Events (Timeline)"""
        ws = wb.create_sheet("Events", 1)

        headers = [
            "Event ID", "App ID", "Company", "Job Title", "Event Type",
            "Status", "Event Date", "Title", "Description", "Confidence"
        ]

        self._write_headers(ws, headers)

        # Data - filter by user if specified
        query = self.db.query(ApplicationEvent).join(Application)
        if self.user_id:
            query = query.filter(Application.user_id == self.user_id)
        events = query.order_by(ApplicationEvent.event_date.desc()).all()

        for row_idx, event in enumerate(events, start=2):
            ws.cell(row_idx, 1, event.id)
            ws.cell(row_idx, 2, event.application_id)
            ws.cell(row_idx, 3, event.application.company_name)
            ws.cell(row_idx, 4, event.application.job_title)
            ws.cell(row_idx, 5, event.event_type.value)
            ws.cell(row_idx, 6, event.status.value)
            ws.cell(row_idx, 7, event.event_date.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row_idx, 8, event.title or "")
            ws.cell(row_idx, 9, (event.description or "")[:200])
            ws.cell(row_idx, 10, f"{event.confidence:.2f}" if event.confidence else "")

        self._autosize_columns(ws)

    def _create_action_queue_sheet(self, wb):
        """Sheet 3: Action Queue"""
        ws = wb.create_sheet("Action Queue", 2)

        headers = [
            "App ID", "Company", "Job Title", "Action Type",
            "Deadline", "Description", "Days Until Deadline"
        ]

        self._write_headers(ws, headers)

        # Data - only applications with action_required, filter by user if specified
        query = self.db.query(Application).filter(Application.action_required == True)
        if self.user_id:
            query = query.filter(Application.user_id == self.user_id)
        applications = query.order_by(Application.action_deadline).all()

        for row_idx, app in enumerate(applications, start=2):
            days_until = None
            if app.action_deadline:
                delta = app.action_deadline - datetime.utcnow()
                days_until = delta.days

            ws.cell(row_idx, 1, app.id)
            ws.cell(row_idx, 2, app.company_name)
            ws.cell(row_idx, 3, app.job_title)
            ws.cell(row_idx, 4, app.action_type.value if app.action_type else "")
            ws.cell(row_idx, 5, app.action_deadline.strftime("%Y-%m-%d") if app.action_deadline else "")
            ws.cell(row_idx, 6, app.action_description or "")
            ws.cell(row_idx, 7, days_until if days_until is not None else "")

            # Highlight overdue
            if days_until is not None and days_until < 0:
                for col in range(1, 8):
                    ws.cell(row_idx, col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        self._autosize_columns(ws)

    def _write_headers(self, ws, headers):
        """Write and style header row"""
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx, header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _autosize_columns(self, ws):
        """Auto-size columns based on content"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _log_export(self, file_path: str, wb):
        """Log export to database"""
        import os

        apps_count = len(list(wb["Applications"].rows)) - 1
        events_count = len(list(wb["Events"].rows)) - 1
        actions_count = len(list(wb["Action Queue"].rows)) - 1

        log = ExportLog(
            file_path=file_path,
            file_size_bytes=os.path.getsize(file_path) if os.path.exists(file_path) else 0,
            applications_count=apps_count,
            events_count=events_count,
            actions_count=actions_count,
            success=True
        )

        self.db.add(log)
        self.db.commit()


# Convenience function
def export_to_excel(db: Session, output_path: str = "job_applications.xlsx", user_id: int = None) -> str:
    """
    Export all data to Excel.

    Usage:
        from app.database import SessionLocal
        from app.excel_exporter import export_to_excel

        db = SessionLocal()
        file_path = export_to_excel(db, "my_applications.xlsx", user_id=1)
        print(f"Exported to: {file_path}")
    """
    exporter = ExcelExporter(db, user_id=user_id)
    return exporter.export_all(output_path)
